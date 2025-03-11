import os
import re
import requests
import openpyxl
from datetime import datetime
from bs4 import BeautifulSoup


def get_html(city: str) -> str:
    METEOPROG_URL = "https://www.meteoprog.com/ua/weather/"
    url = f"{METEOPROG_URL}/{city.title()}" 
    responce = requests.get(url) 
    if responce.status_code == 404:
        return ""
    return responce.text


def get_soup(html: str) -> BeautifulSoup:
    soup = BeautifulSoup(html, "lxml")
    return soup


def parse_temp(s: str) -> str:
    TEMP_REGEX = r"\+?(-?[0-9]*)°"
    temp = re.findall(TEMP_REGEX, s)
    return temp[0] if len(temp) else ""


def get_cur_temp(soup: BeautifulSoup) -> str:
    cur_temp_span = soup.find_all("span", {"dir": "ltr"}, limit=1)[0]
    cur_temp = parse_temp(cur_temp_span.text)
    return cur_temp


def resize_list(lst: list, to_size: int, value) -> list:
    if to_size < len(lst):
        return lst[:to_size]
    lst.extend([value] * (to_size - len(lst)))
    return lst


def zip_lists(lst1: list, lst2: list) -> list:
    return [item for pair in zip(lst1, lst2) for item in pair]


def get_h4_text(tag) -> str:
    h4 = tag.find("h4")
    return h4.text if h4 else "" 


def get_min_max_temp(soup: BeautifulSoup) -> list[str]:
    max_temp_div = soup.find_all("div", {"class": "temperature-max"})
    min_temp_div = soup.find_all("div", {"class": "temperature-min"})

    max_temp_val = list(map(get_h4_text, max_temp_div))
    min_temp_val = list(map(get_h4_text, min_temp_div))

    max_temps = list(map(parse_temp, max_temp_val))
    min_temps = list(map(parse_temp, min_temp_val))

    temps = zip_lists(max_temps, min_temps)
    return resize_list(temps, 6 * 2, "")


def get_temp(city: str) -> list[str]:
    html = get_html(city)
    soup = get_soup(html)
    cur_temp = get_cur_temp(soup)
    five_days_temp  = get_min_max_temp(soup)
    return [cur_temp, *five_days_temp]


def update_xlsx(filename: str, city: str, data: list):
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()

    if city not in wb.sheetnames:
        ws = wb.create_sheet(city)
    else:
        ws = wb[city]

    ws["A1"] = "Дата"
    ws["B1"] = "Поточна температура"
    ws["C1"] = "Мінімальна температура (День 1)"
    ws["D1"] = "Максимальна температура (День 1)"
    ws["E1"] = "Мінімальна температура (День 2)"
    ws["H1"] = "Мінімальна температура (День 2)"
    ws["I1"] = "Мінімальна температура (День 3)"
    ws["J1"] = "Мінімальна температура (День 3)"
    ws["K1"] = "Максимальна температура (День 4)"
    ws["L1"] = "Максимальна температура (День 4)"
    ws["M1"] = "Максимальна температура (День 5)"
    ws["N1"] = "Максимальна температура (День 5)"

    ws.append(data)

    wb.save(filename)
    wb.close()


def get_cur_date() -> str:
    return datetime.today().strftime('%Y-%m-%d')


def format_city_name(city: str) -> str:
    return city.title()


def save_synoptic_stats(city: str, filename: str = "temp_stat_meteoprog.xlsx"):
    cur_date = get_cur_date()   
    data = get_temp(city)
    city = format_city_name(city)

    update_xlsx(filename, city, [cur_date, *data])
    

if __name__ == "__main__":
    save_synoptic_stats("Kyiv")
    save_synoptic_stats("Lviv")
    save_synoptic_stats("Chernigiv")
