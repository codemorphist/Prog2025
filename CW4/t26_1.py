import os
import re
import requests
import openpyxl
from datetime import datetime


def get_html(city: str) -> str:
    SYNOPTIC_URL = "https://sinoptik.ua"
    url = f"{SYNOPTIC_URL}/погода-{city.lower()}" 
    responce = requests.get(url) 
    return responce.text


def get_cur_temp(html: str) -> str:
    CUR_TEMP_REGEX = r"<p class=\"R1ENpvZz\">\+?(\-?[0-9]*).*?</p>"
    cur_temp = re.findall(CUR_TEMP_REGEX, html)
 
    if cur_temp:
        return cur_temp[0]
    else:
        return ""


def resize_list(lst: list, to_size: int, value) -> list:
    if to_size < len(lst):
        return lst[:to_size]
    return lst.extend([value] * (to_size - len(lst)))

    
def get_min_max_temp(html: str) -> list[str]:
    MIN_MAX_TEMP_REGEX = r"<div class=\"\+Ncy59Ya\">.*?<p>\+?(\-?[0-9]*).*?<\/p>"
    temps = re.findall(MIN_MAX_TEMP_REGEX, html)

    return resize_list(temps, 6 * 2, "")


def parse_temp(city: str) -> list[str]:
    html = get_html(city)
    cur_temp = get_cur_temp(html)
    five_days_temp  = get_min_max_temp(html)

    return [cur_temp, *five_days_temp]


def update_xlsx(filename: str, city: str, data: list):
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()

    if city not in wb.sheetnames:
        ws = wb.create_sheet(city)
    else:
        ws = wb[city]

    ws["A1"] = "Дата"
    ws["B1"] = "Поточна температура"
    ws["C1"] = "Мінімальна температура (День 1)"
    ws["D1"] = "Максимальна температура (День 1)"
    ws["E1"] = "Мінімальна температура (День 2)"
    ws["H1"] = "Мінімальна температура (День 2)"
    ws["I1"] = "Мінімальна температура (День 3)"
    ws["J1"] = "Мінімальна температура (День 3)"
    ws["K1"] = "Максимальна температура (День 4)"
    ws["L1"] = "Максимальна температура (День 4)"
    ws["M1"] = "Максимальна температура (День 5)"
    ws["N1"] = "Максимальна температура (День 5)"

    ws.append(data)

    wb.save(filename)
    wb.close()


def get_cur_date() -> str:
    return datetime.today().strftime('%Y-%m-%d')


def format_city_name(city: str) -> str:
    return city.title()


def save_synoptic_stats(city: str, filename: str = "temp_stat.xlsx"):
    cur_date = get_cur_date()   
    data = parse_temp(city)
    city = format_city_name(city)

    update_xlsx(filename, city, [cur_date, *data])
    

if __name__ == "__main__":
    save_synoptic_stats("киев")
    save_synoptic_stats("львов")
    save_synoptic_stats("чернигов")
