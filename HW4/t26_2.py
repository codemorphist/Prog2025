import os
import re
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Color
from datetime import datetime


def get_html(city: str) -> str:
    SYNOPTIC_URL = "https://sinoptik.ua"
    url = f"{SYNOPTIC_URL}/погода-{city.lower()}" 
    responce = requests.get(url) 
    return responce.text


def get_cur_weather(html: str) -> str:
    WEATHER_REGEX = r"<div class=\"bSOXy2ra N5FMbQtj\" aria-label=\"(.*?)\"></div>"
    weather = re.findall(WEATHER_REGEX, html)
    return weather[0] if weather else ""


def get_cur_temp(html: str) -> str:
    CUR_TEMP_REGEX = r"<p class=\"R1ENpvZz\">\+?(\-?[0-9]*).*?</p>"
    cur_temp = re.findall(CUR_TEMP_REGEX, html)
 
    if cur_temp:
        return cur_temp[0]
    else:
        return ""


def resize_list(lst: list, to_size: int, value) -> list:
    if to_size < len(lst):
        return lst[:to_size]
    lst.extend([value] * (to_size - len(lst)))
    return lst

    
def get_min_max_temp(html: str) -> list[str]:
    MIN_MAX_TEMP_REGEX = r"<div class=\"\+Ncy59Ya\">.*?<p>\+?(\-?[0-9]*).*?<\/p>"
    temps = re.findall(MIN_MAX_TEMP_REGEX, html)

    return resize_list(temps, 5 * 2, "")


def parse_data(city: str) -> list[str]:
    html = get_html(city)
    cur_temp = get_cur_temp(html)
    cur_weather = get_cur_weather(html)
    five_days_temp  = get_min_max_temp(html)

    return [cur_temp, cur_weather, *five_days_temp]


def config_ws(ws):
    ws["A1"] = "Дата"
    ws["B1"] = "Поточна температура"
    ws["C1"] = "Поточна погода"
    ws["D1"] = "Мін. темп. (День 1)"
    ws["E1"] = "Макс. темп. (День 1)"
    ws["F1"] = "Мін. темп. (День 2)"
    ws["G1"] = "Макс. темп. (День 2)"
    ws["H1"] = "Мін. темп. (День 3)"
    ws["I1"] = "Макс. темп. (День 3)"
    ws["J1"] = "Мін.. темп. (День 4)"
    ws["K1"] = "Макс. темп. (День 4)"
    ws["L1"] = "Мін. темп. (День 5)"
    ws["M1"] = "Макс. темп. (День 5)"

    fill = PatternFill(start_color="ffd966", end_color="ffd966", 
                       fill_type="solid")
    font = Font(size=12, bold=True)
    border = Border(left=Side(style="thin"), 
                    right=Side(style="thin"), 
                    top=Side(style="thin"), 
                    bottom=Side(style="thin"))

    for c in "ABCDEFGHIJKLM": 
        cell = ws[f"{c}1"]
        cell.fill = fill  
        cell.font = font
        cell.border = border

        ws.column_dimensions[c].width = len(cell.value) + 10


def update_xlsx(filename: str, city: str, data: list):
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()

    if city not in wb.sheetnames:
        ws = wb.create_sheet(city)
    else:
        ws = wb[city]

    config_ws(ws)

    ws.append(data)

    wb.save(filename)
    wb.close()


def get_cur_date() -> str:
    return datetime.today().strftime('%Y-%m-%d')


def format_city_name(city: str) -> str:
    return city.title()


def save_synoptic_stats(city: str, filename: str = "temp_stat_synoptik.xlsx"):
    cur_date = get_cur_date()   
    data = parse_data(city)
    city = format_city_name(city)

    update_xlsx(filename, city, [cur_date, *data])
    

if __name__ == "__main__":
    save_synoptic_stats("киев")
    save_synoptic_stats("львов")
    save_synoptic_stats("чернигов")
